"""
Gera o painel orçamentário UFCG - Modo Escuro, layout Power BI.
Execute sempre que atualizar a planilha, ou deixe o agendador rodar às 9h.
"""
import pandas as pd
import json, os, base64, urllib.request
from datetime import datetime

BASE     = os.path.dirname(os.path.abspath(__file__))
EXCEL    = os.path.join(BASE, 'CONTROLE ORÇAMENTÁRIO  UFCG - PAINEL BI.xlsx')
LOGO     = os.path.join(BASE, 'ufcg_logo.png')
HTML_OUT = os.path.join(BASE, 'painel_orcamentario.html')

# Planilha de planejamento (Limite 2026 = 75% do gasto de 2025), uma aba por unidade.
# Fica junto do script (versionada no repositório) pra também estar disponível
# na atualização automática via GitHub Actions, não só localmente.
EXCEL_PLAN = os.path.join(BASE, '(Pró-Reitorias) PLANEJAMENTO ORÇAMENTÁRIO 2026.xlsx')
PLAN_SHEET_POR_UGR = {
    'PREFEITURA UNIVERSITARIA DA UFCG':             'PU',
    'PRO-REITORIA DE ASSUNTOS COMUNITARIOS DA UFCG': 'PRAC',
    'PRO-REITORIA DE ENSINO DA UFCG':               'PRE',
    'PRO-REITORIA DE EXTENSAO DA UFCG':             'PROPEX',
    'PRO-REITORIA DE GESTAO ADM-FINANCEIRA DA UFCG': 'PRGAF',
    'PRO-REITORIA DE POS-GRADUACAO DA UFCG':        'PRPG',
    'REITORIA/GABINETE DA UFCG':                    'REITORIAGABINETE',
    'SECRETARIA DE PLANEJ. E ORCAMENTO DA UFCG':    'SEPLAN',
    'SECRETARIA DE PROJETOS ESTRATEGICOS DA UFCG':  'SEPE',
    'SECRETARIA DE RECURSOS HUMANOS DA UFCG':       'SRH',
    'SECRETARIA DOS ORGAOS DEL. SUPERIORES DA UFCG': 'SODS',
}
LIMITE_2026 = {}
if os.path.exists(EXCEL_PLAN):
    print(f"[{datetime.now():%Y-%m-%d %H:%M}] Lendo limites 2026 (planejamento Pró-Reitorias)...")
    import openpyxl
    wb_plan = openpyxl.load_workbook(EXCEL_PLAN, data_only=True, read_only=True)
    for ugr_nome, aba in PLAN_SHEET_POR_UGR.items():
        if aba not in wb_plan.sheetnames:
            continue
        ws_plan = wb_plan[aba]
        for row in ws_plan.iter_rows(min_row=1, max_row=8, max_col=4):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().startswith('Limite'):
                    val = ws_plan.cell(row=cell.row, column=cell.column + 1).value
                    if isinstance(val, (int, float)):
                        LIMITE_2026[ugr_nome] = round(float(val), 2)
                    break
    wb_plan.close()
    print(f"  Limites carregados: {len(LIMITE_2026)}/{len(PLAN_SHEET_POR_UGR)} unidades")
else:
    print(f"[{datetime.now():%Y-%m-%d %H:%M}] Planilha de planejamento não encontrada — painel sem coluna Limite 2026.")

# Overrides manuais do Limite 2026, só pro painel (não altera a planilha de
# planejamento original). Pedido em 09/08/2026: PRE passa a usar R$ 36.000,00.
LIMITE_2026_OVERRIDE = {
    'PRO-REITORIA DE ENSINO DA UFCG': 36000.00,
}
LIMITE_2026.update(LIMITE_2026_OVERRIDE)

# Baixar planilha do Google Drive quando rodando em CI (GitHub Actions)
GDRIVE_ID = '13ItIuMTXQllt6y-w9JjoXG26UsB1m0ym'
if os.getenv('CI'):
    print(f"[{datetime.now():%Y-%m-%d %H:%M}] Baixando planilha do Google Drive...")
    import gdown
    gdown.download(id=GDRIVE_ID, output=EXCEL, quiet=False)
    print(f"[{datetime.now():%Y-%m-%d %H:%M}] Planilha baixada.")

# UGRs com siglas e nomes sem "da UFCG"
UGR_META = {
    'PREFEITURA UNIVERSITARIA DA UFCG':            ('Prefeitura Universitária',                       'PU'),
    'PRO-REITORIA DE ASSUNTOS COMUNITARIOS DA UFCG':('Pró-reitoria de Assuntos Comunitários',         'PRAC'),
    'PRO-REITORIA DE ENSINO DA UFCG':              ('Pró-reitoria de Ensino',                         'PRE'),
    'PRO-REITORIA DE EXTENSAO DA UFCG':            ('Pró-reitoria de Extensão',                       'PROPEX'),
    'PRO-REITORIA DE GESTAO ADM-FINANCEIRA DA UFCG':('Pró-reitoria de Gestão Adm-Financeira',         'PRGAF'),
    'PRO-REITORIA DE GESTAO DE PESSOAS':           ('Pró-reitoria de Gestão de Pessoas',              'SRH'),
    'PRO-REITORIA DE POS-GRADUACAO DA UFCG':       ('Pró-reitoria de Pós-Graduação',                  'PRPG'),
    'REITORIA/GABINETE DA UFCG':                   ('Reitoria/Gabinete',                              ''),
    'SECRETARIA DE PLANEJ. E ORCAMENTO DA UFCG':   ('Secretaria de Planej. e Orçamento',              'SEPLAN'),
    'SECRETARIA DE PROJETOS ESTRATEGICOS DA UFCG': ('Secretaria de Projetos Estratégicos',            'SEPE'),
    'SECRETARIA DE RECURSOS HUMANOS DA UFCG':      ('Secretaria de Recursos Humanos',                 'SRH'),
    'SECRETARIA DOS ORGAOS DEL. SUPERIORES DA UFCG':('Secretaria dos Órgãos Del. Superiores',         'SODS'),
}
UGR_ORDER = list(UGR_META.keys())
UGR_ADMIN = set(UGR_META.keys())

print(f"[{datetime.now():%Y-%m-%d %H:%M}] Lendo planilha...")
df = pd.read_excel(EXCEL, sheet_name=0, header=None, skiprows=[1])
df.columns = [
    'Exercicio','Cat_Econ','Acao_Gov','Acao_Gov_Desc','PI','PI_Desc',
    'Nat_Desp','Nat_Desp_Desc','Nat_Desp_Det','Nat_Desp_Det_Desc','PTRES',
    'Fonte_Rec','Fonte_Rec_Desc','Plano_Orc','c14','c15','c16','c17','c18','c19',
    'Prog_Gov','UGR','UGR_Desc','UG_Exec','UG_Exec_Desc',
    'NE_Desc','DOT_ATUAL','DETALHADO','EMPENHADO','DISPONIVEL','LIQUIDADO','PAGOS','RESTOS_PAGAR'
]
FIN = ['DETALHADO','EMPENHADO','LIQUIDADO','PAGOS','DISPONIVEL']
for c in FIN + ['DOT_ATUAL','RESTOS_PAGAR']:
    df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
df['Exercicio'] = pd.to_numeric(df['Exercicio'], errors='coerce')
df = df.dropna(subset=['Exercicio'])
df['Exercicio'] = df['Exercicio'].astype(int)

for col in ['UGR_Desc','Nat_Desp','Nat_Desp_Desc','Nat_Desp_Det','Nat_Desp_Det_Desc','PI_Desc','PI']:
    df[col] = df[col].fillna('').astype(str).str.strip()
df['UGR_Nome'] = df['UGR_Desc'].where(df['UGR_Desc'] != '', df['UGR'].astype(str)).str.strip().str.upper()
df['Nat_Desp_Desc'] = df['Nat_Desp_Desc'].replace({'nan':'','Natureza Despesa':''}).str.strip()

# PIs de "GESTAO DA UNIDADE-X" que a planilha às vezes lança sob a UGR genérica
# 159195 (UNIVERSIDADE FEDERAL DE CAMPINA GRANDE) em vez da UGR real da unidade.
# Cada um desses PIs pertence a uma única unidade (conferido na planilha de
# referência "PIs" do Planejamento Orçamentário 2026) — reclassifica pela UGR certa.
PI_UGR_FORCA = {
    'M20RKG01G1N':  'PRO-REITORIA DE ASSUNTOS COMUNITARIOS DA UFCG',  # GESTAO DA UNIDADE-PRAC
    'M20RKG01A1N':  'REITORIA/GABINETE DA UFCG',                      # GESTAO DA UNIDADE-REITORIA/GABINETE
    'O20RKO01E1N':  'PRO-REITORIA DE POS-GRADUACAO DA UFCG',          # GESTAO DA UNIDADE-PRPG
    'M20RKG01H1N':  'SECRETARIA DE RECURSOS HUMANOS DA UFCG',         # GESTAO DA UNIDADE-SRH
    'M20RKG01C1N':  'PRO-REITORIA DE GESTAO ADM-FINANCEIRA DA UFCG',  # GESTAO DA UNIDADE-PRA
    'M20RKG01A0N':  'SECRETARIA DOS ORGAOS DEL. SUPERIORES DA UFCG',  # GESTAO DA UNIDADE-SODS
    'M20RKG01D1N':  'PRO-REITORIA DE ENSINO DA UFCG',                 # GESTAO DA UNIDADE-PRE
    'M20RKG01I1N':  'PREFEITURA UNIVERSITARIA DA UFCG',               # GESTAO DA UNIDADE-PU
}
for pi_code, ugr_nome in PI_UGR_FORCA.items():
    df.loc[df['PI'].str.upper() == pi_code, 'UGR_Nome'] = ugr_nome

# PI "de gestão" de cada UGR — o único que tem Limite 2026 definido na planilha
# de planejamento. Usado pra separar, dentro da UGR, o que é comparável ao
# limite planejado do que é "fora do planejamento" (outros PIs sem limite).
GESTAO_PI_POR_UGR = {
    'PRO-REITORIA DE ASSUNTOS COMUNITARIOS DA UFCG':  'M20RKG01G1N',
    'REITORIA/GABINETE DA UFCG':                      'M20RKG01A1N',
    'PRO-REITORIA DE POS-GRADUACAO DA UFCG':          'O20RKO01E1N',
    'SECRETARIA DE RECURSOS HUMANOS DA UFCG':         'M20RKG01H1N',
    'PRO-REITORIA DE GESTAO ADM-FINANCEIRA DA UFCG':  'M20RKG01C1N',
    'SECRETARIA DOS ORGAOS DEL. SUPERIORES DA UFCG':  'M20RKG01A0N',
    'PRO-REITORIA DE ENSINO DA UFCG':                 'M20RKG01D1N',
    'PREFEITURA UNIVERSITARIA DA UFCG':               'M20RKG01I1N',
    'PRO-REITORIA DE EXTENSAO DA UFCG':               'M20RKG01F1N',
    'SECRETARIA DE PLANEJ. E ORCAMENTO DA UFCG':      'M20RKG01B1N',
    'SECRETARIA DE PROJETOS ESTRATEGICOS DA UFCG':    'M20RKG01B3N',
}

# NDs que sempre contam como "Fora do planejamento", mesmo quando lançadas
# sob o PI de gestão — não fazem parte do que o Limite 2026 cobre.
ND_SEMPRE_FORA_PLANO = set()

# ── Filtros fixos (conforme Power BI) ───────────────────────────────────────
df = df[df['Acao_Gov'].str.upper() == '20RK']
PI_EXCLUIR = {'M20RKG01IJN', 'M20RKG01ILN', 'M20RKG01IIN'}
df = df[~df['PI'].str.upper().isin(PI_EXCLUIR)]
# ─────────────────────────────────────────────────────────────────────────────

anos = sorted(df['Exercicio'].unique().tolist())
ugrs = [u for u in UGR_ORDER if u in df['UGR_Nome'].values]

# NDs disponíveis (para filtro)
df_admin = df[df['UGR_Nome'].isin(UGR_ADMIN)]
nds_set = sorted(
    nd for nd in df_admin['Nat_Desp_Desc'].unique()
    if nd and nd not in ('nan','','Natureza Despesa')
)

def s(v): return round(float(v), 2)

def saldo(det, liq): return round(det - liq, 2)

def nd_label(nd):
    if not nd or nd in ('nan','') or nd.strip().upper() == 'NAO SE APLICA':
        return 'NAO SE APLICA (VALOR DISPONÍVEL PARA EMPENHAR)'
    return nd

def build_nd_tree(du):
    """Constrói a árvore ND -> NDD -> PI para um subconjunto de linhas."""
    nds = []
    for nd_raw, gnd in du.groupby('Nat_Desp_Desc'):
        nd_desc = nd_label(nd_raw)
        if gnd['DETALHADO'].sum() == 0 and gnd['EMPENHADO'].sum() == 0 and gnd['LIQUIDADO'].sum() == 0 and gnd['PAGOS'].sum() == 0:
            continue
        _nd_det = s(gnd['DETALHADO'].sum()); _nd_liq = s(gnd['LIQUIDADO'].sum())
        nd_row = {
            'label': nd_desc, 'nd_raw': nd_raw,
            'DET': _nd_det, 'EMP': s(gnd['EMPENHADO'].sum()),
            'LIQ': _nd_liq, 'PAG': s(gnd['PAGOS'].sum()),
            'DISP': saldo(_nd_det, _nd_liq), 'children': []
        }
        for ndd_raw, gndd in gnd.groupby('Nat_Desp_Det_Desc'):
            ndd_strip = ndd_raw.strip()
            if not ndd_strip or ndd_raw in ('nan','') or ndd_strip.upper() == 'NAO SE APLICA':
                ndd_desc = 'NAO SE APLICA (VALOR DISPONÍVEL PARA EMPENHAR)'
            else:
                ndd_desc = ndd_strip
            if gndd['DETALHADO'].sum() == 0 and gndd['EMPENHADO'].sum() == 0 and gndd['LIQUIDADO'].sum() == 0 and gndd['PAGOS'].sum() == 0:
                continue
            _ndd_det = s(gndd['DETALHADO'].sum()); _ndd_liq = s(gndd['LIQUIDADO'].sum())
            ndd_row = {
                'label': ndd_desc,
                'DET': _ndd_det, 'EMP': s(gndd['EMPENHADO'].sum()),
                'LIQ': _ndd_liq, 'PAG': s(gndd['PAGOS'].sum()),
                'DISP': saldo(_ndd_det, _ndd_liq), 'children': []
            }
            for pi_val, gpi in gndd.groupby('PI'):
                if pi_val in ('','nan','PI'): continue
                if gpi['DETALHADO'].sum() == 0 and gpi['EMPENHADO'].sum() == 0 and gpi['LIQUIDADO'].sum() == 0 and gpi['PAGOS'].sum() == 0:
                    continue
                pi_desc = gpi['PI_Desc'].iloc[0] if not gpi['PI_Desc'].empty else ''
                lbl = pi_val
                if pi_desc and pi_desc not in ('nan',''):
                    lbl += ' — ' + pi_desc[:55]
                _pi_det = s(gpi['DETALHADO'].sum()); _pi_liq = s(gpi['LIQUIDADO'].sum())
                ndd_row['children'].append({
                    'label': lbl,
                    'DET': _pi_det, 'EMP': s(gpi['EMPENHADO'].sum()),
                    'LIQ': _pi_liq, 'PAG': s(gpi['PAGOS'].sum()),
                    'DISP': saldo(_pi_det, _pi_liq)
                })
            nd_row['children'].append(ndd_row)
        nds.append(nd_row)
    return nds

print("Agregando dados hierárquicos...")
dados = {}
for ano in anos:
    da = df[df['Exercicio'] == ano]
    dados[str(ano)] = {}
    for ugr in ugrs:
        du = da[da['UGR_Nome'] == ugr]
        if du.empty or du['DETALHADO'].sum() == 0:
            continue
        _det = s(du['DETALHADO'].sum()); _liq = s(du['LIQUIDADO'].sum())
        ugr_row = {
            'DET': _det, 'EMP': s(du['EMPENHADO'].sum()),
            'LIQ': _liq, 'PAG': s(du['PAGOS'].sum()),
            'DISP': saldo(_det, _liq), 'children': []
        }
        if ano == 2026 and ugr in LIMITE_2026:
            ugr_row['LIMITE'] = LIMITE_2026[ugr]

        gestao_pi = GESTAO_PI_POR_UGR.get(ugr)
        if gestao_pi:
            eh_pi_gestao = du['PI'].str.upper() == gestao_pi
            eh_nd_fora = du['Nat_Desp_Desc'].str.upper().isin(ND_SEMPRE_FORA_PLANO)
            dg = du[eh_pi_gestao & ~eh_nd_fora]
            do_ = du[~eh_pi_gestao | eh_nd_fora]
            _g_det = s(dg['DETALHADO'].sum()); _g_liq = s(dg['LIQUIDADO'].sum())
            _f_det = s(do_['DETALHADO'].sum()); _f_liq = s(do_['LIQUIDADO'].sum())
            gestao_node = {
                'label': 'PLANEJADO', 'is_group': True,
                'DET': _g_det, 'EMP': s(dg['EMPENHADO'].sum()),
                'LIQ': _g_liq, 'PAG': s(dg['PAGOS'].sum()),
                'DISP': saldo(_g_det, _g_liq), 'children': build_nd_tree(dg)
            }
            fora_node = {
                'label': 'Fora do planejamento (demais PIs, sem limite definido)', 'is_group': True,
                'DET': _f_det, 'EMP': s(do_['EMPENHADO'].sum()),
                'LIQ': _f_liq, 'PAG': s(do_['PAGOS'].sum()),
                'DISP': saldo(_f_det, _f_liq), 'children': build_nd_tree(do_)
            }
            ugr_row['children'] = [gestao_node] if _f_det == 0 and _f_liq == 0 and fora_node['EMP'] == 0 and fora_node['PAG'] == 0 else [gestao_node, fora_node]
            if ano == 2026 and ugr in LIMITE_2026:
                ugr_row['GESTAO'] = {'DET': _g_det, 'EMP': gestao_node['EMP'], 'LIQ': _g_liq, 'PAG': gestao_node['PAG']}
                ugr_row['FORA_PLANO'] = {'DET': _f_det, 'EMP': fora_node['EMP'], 'LIQ': _f_liq, 'PAG': fora_node['PAG']}
        else:
            ugr_row['children'] = build_nd_tree(du)

        dados[str(ano)][ugr] = ugr_row

# Logo
logo_b64 = ''
if os.path.exists(LOGO):
    with open(LOGO,'rb') as f:
        logo_b64 = 'data:image/png;base64,' + base64.b64encode(f.read()).decode()

# Serializar metadados de UGR
ugr_meta_js = {k: {'nome': v[0], 'sigla': v[1]} for k, v in UGR_META.items()}

payload = json.dumps({
    'anos': anos, 'ugrs': ugrs, 'ugr_meta': ugr_meta_js,
    'nds': nds_set, 'dados': dados,
    'atualizado': datetime.now().strftime('%d/%m/%Y')
}, ensure_ascii=False)
print(f"Anos: {anos} | UGRs: {len(ugrs)} | NDs: {len(nds_set)}")

HTML = f"""<!DOCTYPE html>
<html lang="pt-BR" data-theme="light">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Monitoramento da Execução Orçamentária - UFCG</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
:root{{
  --bg:#1a1f2e;--bg2:#0d1526;--bg3:#1e2740;--border:#2a3555;
  --text:#d0d7e8;--text2:#8a96b8;--text3:#4b5675;
  --accent:#4a8bc4;--green:#2ea85a;
  --c-det:#5c8fe8;--c-emp:#d4a017;--c-liq:#c47a35;--c-pag:#3a9e5f;
  --bar-det:#003D7C;--bar-emp:#d4a017;--bar-liq:#c47a35;--bar-pag:#2d8f52;
  --ugr-bg:#1e2740;--ugr-text:#c8d0e8;
}}
[data-theme=light]{{
  --bg:#f0f4f8;--bg2:#003D7C;--bg3:#e2eaf3;--border:#c5d3e0;
  --text:#1a2740;--text2:#4a5a7a;--text3:#8a96b8;
  --accent:#003D7C;--green:#00843D;
  --c-det:#003D7C;--c-emp:#8a5f00;--c-liq:#8a4a1a;--c-pag:#00843D;
  --bar-det:#003D7C;--bar-emp:#d4a017;--bar-liq:#c47a35;--bar-pag:#2d8f52;
  --ugr-bg:#dce8f5;--ugr-text:#003D7C;
}}
body{{font-family:'Segoe UI',Arial,sans-serif;background:var(--bg);color:var(--text);font-size:13px}}

/* HEADER */
.hdr{{background:var(--bg2);padding:10px 20px;display:flex;align-items:center;gap:12px;border-bottom:2px solid var(--border)}}
.hdr img{{width:54px;height:54px;object-fit:contain}}
.hdr-logo-txt{{width:54px;height:54px;background:#003D7C;border-radius:50%;display:flex;align-items:center;justify-content:center;font-weight:900;font-size:12px;color:white;flex-shrink:0}}
.hdr h1{{font-size:17px;font-weight:700;color:#e8edf8;letter-spacing:.3px}}
[data-theme=light] .hdr h1{{color:white}}
.hdr p{{font-size:10px;color:#6b7a99;margin-top:2px}}
[data-theme=light] .hdr p{{color:#b8d0f0}}
.hdr-actions{{margin-left:auto;display:flex;gap:8px;align-items:center;flex-shrink:0}}
.btn-theme{{background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.25);color:white;border-radius:6px;padding:5px 12px;cursor:pointer;font-size:11px}}
.btn-theme:hover{{background:rgba(255,255,255,.22)}}

/* LAYOUT */
.body{{display:flex;height:calc(100vh - 126px)}}
.body.sb-hidden .sidebar{{max-width:0;width:0;padding:0;overflow:hidden;border:none}}
.body.sb-hidden .main{{flex:1}}

/* SIDEBAR TOGGLE TAB */
.sb-toggle{{width:14px;flex-shrink:0;background:var(--bg3);border:none;
  border-left:1px solid var(--border);border-right:1px solid var(--border);
  color:var(--text3);cursor:pointer;font-size:9px;display:flex;
  align-items:center;justify-content:center;padding:0;transition:background .15s}}
.sb-toggle:hover{{background:var(--accent);color:white}}

/* SIDEBAR */
.sidebar{{width:265px;max-width:265px;flex-shrink:0;background:var(--bg2);border-right:1px solid var(--border);padding:14px;overflow-y:auto;display:flex;flex-direction:column;gap:14px}}
[data-theme=light] .sidebar{{background:#f8fafc}}
.sb-title{{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.8px;color:var(--accent);margin-bottom:7px}}
.search-box{{display:flex;align-items:center;gap:6px;background:var(--bg3);border:1px solid var(--border);border-radius:6px;padding:5px 9px}}
.search-box svg{{flex-shrink:0}}
.search-box input{{background:none;border:none;outline:none;color:var(--text);font-size:12px;width:100%}}
.search-box input::placeholder{{color:var(--text3)}}
.ugr-list{{list-style:none;display:flex;flex-direction:column;gap:1px;max-height:260px;overflow-y:auto;margin-top:7px}}
.ugr-list li{{display:flex;align-items:flex-start;gap:7px;padding:5px 6px;border-radius:5px;cursor:pointer;font-size:11.5px;color:var(--text2);line-height:1.3}}
.ugr-list li:hover{{background:var(--bg3);color:var(--text)}}
.ugr-list li.active{{background:var(--ugr-bg);color:var(--accent);font-weight:600}}
.ugr-list li input[type=radio]{{accent-color:var(--accent);margin-top:3px;flex-shrink:0}}
.nd-list{{list-style:none;display:flex;flex-direction:column;gap:1px;max-height:180px;overflow-y:auto;margin-top:7px}}
.nd-list li{{display:flex;align-items:flex-start;gap:6px;padding:4px 5px;border-radius:4px;cursor:pointer;font-size:11px;color:var(--text2);line-height:1.3}}
.nd-list li:hover{{background:var(--bg3);color:var(--text)}}
.nd-list li.active{{background:var(--ugr-bg);color:var(--accent);font-weight:600}}
.nd-list li input[type=checkbox]{{accent-color:var(--accent);margin-top:2px;flex-shrink:0}}
.btn-limpar{{width:100%;margin-top:6px;padding:5px;background:var(--bg3);border:1px solid var(--border);color:var(--text2);border-radius:5px;cursor:pointer;font-size:11px}}
.btn-limpar:hover{{color:var(--text)}}

/* MAIN */
.main{{flex:1;overflow-y:auto;display:flex;flex-direction:column}}

/* ANO BAR + TOOLBAR */
.ano-bar{{background:var(--bg2);border-bottom:1px solid var(--border);padding:7px 16px;display:flex;align-items:center;gap:8px;flex-wrap:wrap}}
[data-theme=light] .ano-bar{{background:white}}
.ano-label{{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.8px;color:var(--accent)}}
.ano-select{{background:var(--bg3);border:1px solid var(--border);color:var(--text);padding:4px 10px;border-radius:6px;font-size:12px;cursor:pointer}}
.toolbar{{margin-left:auto;display:flex;gap:5px;align-items:center;flex-wrap:wrap}}
.tb-btn{{display:flex;align-items:center;gap:4px;padding:4px 9px;border:1px solid var(--border);background:var(--bg3);color:var(--text2);border-radius:5px;cursor:pointer;font-size:11px;white-space:nowrap}}
.tb-btn:hover{{color:var(--text);border-color:var(--accent)}}
.tbl-search{{display:flex;align-items:center;gap:5px;background:var(--bg3);border:1px solid var(--border);border-radius:5px;padding:3px 8px}}
.tbl-search input{{background:none;border:none;outline:none;color:var(--text);font-size:11px;width:110px}}
.tbl-search input::placeholder{{color:var(--text3)}}

.main-search-bar{{display:flex;align-items:center;gap:10px;padding:12px 18px;border-bottom:1px solid var(--border)}}
.main-search-bar input{{flex:1;background:var(--bg3);border:1px solid var(--border);border-radius:6px;padding:9px 14px;color:var(--text);font-size:13px;outline:none}}
.main-search-bar input:focus{{border-color:var(--accent)}}
.main-search-bar input::placeholder{{color:var(--text3)}}
.btn-limpar-busca{{background:#e0503c;border:none;color:white;border-radius:6px;padding:9px 16px;font-size:12px;font-weight:600;cursor:pointer;white-space:nowrap}}
.btn-limpar-busca:hover{{background:#c8432f}}

/* KPI CARDS */
.kpi-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;padding:12px 18px 8px;border-bottom:1px solid var(--border)}}
.kpi-card{{background:var(--bg3);border:1px solid var(--border);border-radius:8px;padding:10px 14px;position:relative;overflow:hidden}}
.kpi-card::before{{content:'';position:absolute;top:0;left:0;width:3px;height:100%;background:var(--kpi-c)}}
.kpi-lbl{{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.6px;color:var(--text3);margin-bottom:4px}}
.kpi-val{{font-size:15px;font-weight:700;color:var(--kpi-c);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.kpi-sub{{font-size:10px;color:var(--text3);margin-top:3px}}
.kc-det{{--kpi-c:var(--c-det)}}
.kc-emp{{--kpi-c:var(--c-emp)}}
.kc-liq{{--kpi-c:var(--c-liq)}}
.kc-disp{{--kpi-c:var(--green)}}

/* EXECUCAO */
.exec-section{{padding:10px 20px 8px;border-bottom:1px solid var(--border)}}
.exec-title{{font-size:11px;font-weight:700;text-transform:uppercase;color:var(--accent);letter-spacing:.6px;margin-bottom:9px}}
.bars{{display:flex;flex-direction:column;gap:6px}}
.bar-row{{display:grid;grid-template-columns:100px 1fr;align-items:center;gap:10px}}
.bar-label{{font-size:11px;font-weight:700;text-transform:uppercase;color:var(--text2);text-align:right}}
.bar-track{{height:30px;background:var(--bg3);border-radius:4px;overflow:hidden}}
.bar-fill{{height:100%;border-radius:4px;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;color:white;white-space:nowrap;padding:0 12px;transition:width .5s ease;min-width:160px}}

/* TABLE */
.tbl-section{{padding:10px 16px;flex:1}}
.tbl-title{{font-size:12px;font-weight:700;text-transform:uppercase;color:var(--accent);letter-spacing:.6px;margin-bottom:8px}}
.tbl-wrap{{overflow-x:auto}}
table{{width:100%;border-collapse:collapse;font-size:11.5px}}
thead th{{
  background:var(--bg2);color:var(--text3);padding:7px 10px;text-align:right;
  font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.4px;
  white-space:nowrap;border-bottom:1px solid var(--border);position:sticky;top:0;z-index:1
}}
[data-theme=light] thead th{{background:#e2eaf3}}
thead th:first-child{{text-align:left;width:280px}}
thead th.th-det{{color:var(--c-det)}}
thead th.th-emp{{color:var(--c-emp)}}
thead th.th-liq{{color:var(--c-liq)}}
thead th.th-pag{{color:var(--c-pag)}}

tbody td{{padding:6px 10px;border-bottom:1px solid var(--bg3);text-align:right;white-space:nowrap;color:var(--text2)}}
[data-theme=light] tbody td{{border-bottom:1px solid var(--border)}}
tbody td:first-child{{text-align:left}}
tbody td.td-det{{color:var(--c-det);font-weight:500}}
tbody td.td-emp{{color:var(--c-emp);font-weight:500}}
tbody td.td-liq{{color:var(--c-liq);font-weight:500}}
tbody td.td-pag{{color:var(--c-pag);font-weight:500}}

tbody tr.ugr-row td{{background:var(--ugr-bg);font-weight:700}}
tbody tr.ugr-row td:first-child{{color:var(--ugr-text)}}
tbody tr.ugr-row td.td-det{{color:var(--c-det)}}
tbody tr.ugr-row td.td-emp{{color:var(--c-emp)}}
tbody tr.ugr-row td.td-liq{{color:var(--c-liq)}}
tbody tr.ugr-row td.td-pag{{color:var(--c-pag)}}
tbody tr.ugr-row:hover td{{filter:brightness(1.08)}}
tbody tr.nd-row:hover td,tbody tr.ndd-row:hover td,tbody tr.pi-row:hover td{{background:var(--bg3)}}
[data-theme=light] tbody tr.nd-row:hover td,[data-theme=light] tbody tr.ndd-row:hover td,[data-theme=light] tbody tr.pi-row:hover td{{background:#e2eaf3}}
tbody tr.ndd-row td{{font-size:11px}}
tbody tr.group-row td{{font-weight:600;background:var(--bg3)}}
[data-theme=light] tbody tr.group-row td{{background:#eef3f8}}
tbody tr.group-row-fora td:first-child{{color:var(--c-liq)}}
.excedente-tag{{font-size:10px;font-weight:700;color:#d13438;background:rgba(209,52,56,.12);border-radius:10px;padding:2px 8px;display:inline-block;vertical-align:middle;margin-left:6px}}
tbody tr.pi-row td{{font-size:10.5px;color:var(--text3)}}
tbody tr.total-row td{{background:var(--bg2);font-weight:700;color:var(--text);border-top:2px solid var(--border)}}
[data-theme=light] tbody tr.total-row td{{background:#dce8f5}}
tbody tr.total-row td.td-det{{color:var(--c-det)}}
tbody tr.total-row td.td-emp{{color:var(--c-emp)}}
tbody tr.total-row td.td-liq{{color:var(--c-liq)}}
tbody tr.total-row td.td-pag{{color:var(--c-pag)}}
.tr-hidden{{display:none}}

.expand-btn{{
  cursor:pointer;user-select:none;display:inline-flex;align-items:center;
  justify-content:center;width:16px;height:16px;border:1px solid var(--border);
  border-radius:3px;margin-right:5px;font-size:10px;color:var(--accent);
  flex-shrink:0;background:var(--bg3);vertical-align:middle
}}
.indent1{{padding-left:24px}}
.indent2{{padding-left:44px}}
.indent3{{padding-left:64px}}

.ndd-cell{{display:inline-block;position:relative;cursor:default}}
.pi-tip{{
  display:none;position:absolute;left:calc(100% + 6px);top:50%;
  transform:translateY(-50%);z-index:100;
  background:rgba(10,20,55,.35);color:#a8c4e8;border-radius:3px;
  padding:1px 6px;font-size:9px;pointer-events:none;white-space:nowrap;
}}
[data-theme=light] .pi-tip{{background:rgba(0,25,70,.22);color:#4a6a9a}}
.ndd-cell:hover .pi-tip{{display:block}}

.exec-badge{{display:inline-block;font-size:9px;font-weight:700;padding:1px 6px;border-radius:10px;margin-left:7px;vertical-align:middle}}
.bg-g{{background:rgba(46,168,90,.15);color:#2ea85a}}
.bg-a{{background:rgba(212,160,23,.15);color:#d4a017}}
.bg-r{{background:rgba(196,122,53,.18);color:#c47a35}}
[data-theme=light] .bg-g{{background:rgba(0,132,61,.12);color:#00843D}}
[data-theme=light] .bg-a{{background:rgba(138,95,0,.12);color:#8a6000}}
[data-theme=light] .bg-r{{background:rgba(138,74,26,.12);color:#8a4a1a}}

.footer{{background:var(--bg2);border-top:1px solid var(--border);font-size:10px;color:var(--text3);padding:6px 20px;text-align:center}}
[data-theme=light] .footer{{background:#e2eaf3}}
.footer span{{color:var(--accent);font-weight:600}}

::-webkit-scrollbar{{width:5px;height:5px}}
::-webkit-scrollbar-track{{background:var(--bg2)}}
::-webkit-scrollbar-thumb{{background:var(--border);border-radius:3px}}

@media print{{
  html,body{{background:white!important;color:#111!important;font-size:11px}}
  .hdr{{background:#003D7C!important;-webkit-print-color-adjust:exact;print-color-adjust:exact;padding:8px 16px}}
  .hdr h1,.hdr p{{color:white!important}}
  .sidebar,.toolbar,.btn-theme,.tb-btn,.tbl-search,.ano-bar .toolbar{{display:none!important}}
  .body{{display:block!important;height:auto!important}}
  .main{{overflow:visible!important}}
  .kpi-grid{{display:grid!important;grid-template-columns:repeat(4,1fr);gap:6px;padding:8px 12px}}
  .kpi-card{{background:#f0f4f8!important;border:1px solid #c5d3e0!important;-webkit-print-color-adjust:exact;print-color-adjust:exact}}
  .kpi-val{{font-size:13px!important}}
  .exec-section{{padding:6px 12px}}
  .bar-track{{background:#e8eef5!important;-webkit-print-color-adjust:exact;print-color-adjust:exact}}
  .bar-fill{{-webkit-print-color-adjust:exact;print-color-adjust:exact;min-width:60px!important;font-size:10px!important}}
  .tbl-section{{padding:6px 12px}}
  .tbl-wrap{{overflow:visible!important}}
  table{{font-size:9px!important;width:100%!important}}
  thead th{{background:#003D7C!important;color:white!important;-webkit-print-color-adjust:exact;print-color-adjust:exact;padding:4px 6px!important}}
  tbody td{{padding:3px 6px!important;border-bottom:1px solid #ddd!important}}
  tbody tr.ugr-row td{{background:#dce8f5!important;-webkit-print-color-adjust:exact;print-color-adjust:exact}}
  tbody tr.total-row td{{background:#e8eef5!important;-webkit-print-color-adjust:exact;print-color-adjust:exact}}
  .exec-badge{{-webkit-print-color-adjust:exact;print-color-adjust:exact}}
  .ano-bar{{background:#f0f4f8!important}}
  .footer{{background:#e8eef5!important;font-size:8px!important}}
  .tr-hidden{{display:none!important}}
  @page{{margin:1cm;size:A4 landscape}}
}}
</style>
<script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
</head>
<body>
<div class="hdr">
  {'<img src="' + logo_b64 + '" alt="UFCG"/>' if logo_b64 else '<div class="hdr-logo-txt">UFCG</div>'}
  <div>
    <h1>MONITORAMENTO DA EXECUÇÃO ORÇAMENTÁRIA - UFCG</h1>
    <p>SEPLAN &middot; Dados extraídos do Tesouro Gerencial / SIAFI &middot; Atualizado em: <span id="dataAtual">—</span></p>
    <p style="margin-top:3px;font-size:10px;">Escopo: Reitoria &middot; Pró-Reitorias &middot; Secretarias &middot; Prefeitura Universitária</p>
  </div>
  <div class="hdr-actions">
    <button class="btn-theme" id="hamburgerBtn" onclick="toggleSidebar()" title="Recolher/expandir filtros">&#9776; Filtros</button>
    <button class="btn-theme" id="themeBtn" onclick="toggleTheme()">🌙 Modo Escuro</button>
  </div>
</div>

<!-- FILTRO GERAL -->
<div class="main-search-bar">
  <input type="text" placeholder="Buscar" id="tblSearch" oninput="update()">
  <button class="btn-limpar-busca" onclick="document.getElementById('tblSearch').value='';update()">&#10005; Limpar</button>
</div>

<div class="body">
  <!-- SIDEBAR -->
  <div class="sidebar">
    <div>
      <div class="sb-title">Unidade Gestora Responsável</div>
      <div class="search-box">
        <svg width="12" height="12" fill="none" stroke="var(--text3)" stroke-width="2" viewBox="0 0 24 24"><circle cx="11" cy="11" r="8"/><path d="m21 21-4.35-4.35"/></svg>
        <input type="text" placeholder="Pesquisar" id="ugrSearch" oninput="filterUGR(this.value)">
      </div>
      <ul class="ugr-list" id="ugrList"></ul>
      <button class="btn-limpar" id="btnLimparUGR" onclick="clearUGR()" style="display:none">Limpar seleção UGR</button>
    </div>
    <div>
      <div class="sb-title">Natureza de Despesa</div>
      <div class="search-box">
        <svg width="12" height="12" fill="none" stroke="var(--text3)" stroke-width="2" viewBox="0 0 24 24"><circle cx="11" cy="11" r="8"/><path d="m21 21-4.35-4.35"/></svg>
        <input type="text" placeholder="Pesquisar" id="ndSearch" oninput="filterND(this.value)">
      </div>
      <ul class="nd-list" id="ndList"></ul>
      <button class="btn-limpar" onclick="clearND()">Limpar seleção ND</button>
    </div>
  </div>

  <!-- SIDEBAR TOGGLE TAB -->
  <button class="sb-toggle" id="sbToggle" onclick="toggleSidebar()" title="Recolher/expandir painel">◀</button>

  <!-- MAIN -->
  <div class="main">
    <!-- ANO BAR + TOOLBAR -->
    <div class="ano-bar">
      <span class="ano-label">ANO</span>
      <select class="ano-select" id="anoSelect" onchange="update()"></select>
      <div class="toolbar">
        <button class="tb-btn" onclick="expandAll()">
          <svg width="11" height="11" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><path d="M8 3H5a2 2 0 0 0-2 2v3m18 0V5a2 2 0 0 0-2-2h-3m0 18h3a2 2 0 0 0 2-2v-3M3 16v3a2 2 0 0 0 2 2h3"/></svg>
          Expandir tudo
        </button>
        <button class="tb-btn" onclick="collapseAll()">
          <svg width="11" height="11" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><path d="M8 3v3a2 2 0 0 1-2 2H3m18 0h-3a2 2 0 0 1-2-2V3m0 18v-3a2 2 0 0 1 2-2h3M3 16h3a2 2 0 0 1 2 2v3"/></svg>
          Recolher tudo
        </button>
        <button class="tb-btn" onclick="exportExcel()">
          <svg width="11" height="11" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><rect x="3" y="3" width="18" height="18" rx="2"/><path d="M3 9h18M9 21V9"/></svg>
          Excel
        </button>
        <button class="tb-btn" onclick="exportPDF()">
          <svg width="11" height="11" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14,2 14,8 20,8"/></svg>
          PDF
        </button>
        <button class="tb-btn" id="copyBtn" onclick="copyLink(this)">
          <svg width="11" height="11" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><path d="M10 13a5 5 0 0 0 7.54.54l3-3a5 5 0 0 0-7.07-7.07l-1.72 1.71"/><path d="M14 11a5 5 0 0 0-7.54-.54l-3 3a5 5 0 0 0 7.07 7.07l1.71-1.71"/></svg>
          Copiar link
        </button>
      </div>
    </div>

    <!-- KPI CARDS -->
    <div class="kpi-grid">
      <div class="kpi-card kc-det">
        <div class="kpi-lbl">Dotado</div>
        <div class="kpi-val" id="kDET">—</div>
        <div class="kpi-sub">100% do orçamento</div>
      </div>
      <div class="kpi-card kc-emp">
        <div class="kpi-lbl">Empenhado</div>
        <div class="kpi-val" id="kEMP">—</div>
        <div class="kpi-sub" id="kEMPs">—</div>
      </div>
      <div class="kpi-card kc-liq">
        <div class="kpi-lbl">Liquidado</div>
        <div class="kpi-val" id="kLIQ">—</div>
        <div class="kpi-sub" id="kLIQs">—</div>
      </div>
      <div class="kpi-card kc-disp">
        <div class="kpi-lbl">Saldo Disponível</div>
        <div class="kpi-val" id="kDISP">—</div>
        <div class="kpi-sub">DET − LIQ</div>
      </div>
    </div>

    <!-- BARRAS DE EXECUÇÃO -->
    <div class="exec-section">
      <div class="exec-title">Execução Orçamentária</div>
      <div class="bars">
        <div class="bar-row">
          <div class="bar-label">Detalhado</div>
          <div class="bar-track"><div class="bar-fill" id="bDET" style="background:var(--bar-det);width:100%">—</div></div>
        </div>
        <div class="bar-row">
          <div class="bar-label">Empenhado</div>
          <div class="bar-track"><div class="bar-fill" id="bEMP" style="background:var(--bar-emp)">—</div></div>
        </div>
        <div class="bar-row">
          <div class="bar-label">Liquidado</div>
          <div class="bar-track"><div class="bar-fill" id="bLIQ" style="background:var(--bar-liq)">—</div></div>
        </div>
        <div class="bar-row">
          <div class="bar-label">Pagos</div>
          <div class="bar-track"><div class="bar-fill" id="bPAG" style="background:var(--bar-pag)">—</div></div>
        </div>
      </div>
    </div>

    <!-- TABLE -->
    <div class="tbl-section">
      <div class="tbl-title">Detalhamento por UGR</div>
      <div class="tbl-wrap">
        <table>
          <thead>
            <tr>
              <th>UGR</th>
              <th title="Limite 2026 = 75% do gasto de 2025 (planejamento Pró-Reitorias)" style="cursor:help">Limite 2026 &#9432;</th>
              <th class="th-det">Detalhado</th>
              <th class="th-emp">Empenhado</th>
              <th class="th-liq">Liquidado</th>
              <th class="th-pag">Pagos</th>
              <th title="Disponível = DETALHADO − LIQUIDADO" style="cursor:help">Disponível (Det. - Liq.) &#9432;</th>
            </tr>
          </thead>
          <tbody id="tbody"></tbody>
        </table>
      </div>
    </div>

    <div class="footer">
      Fonte: Tesouro Gerencial | Ação 20RK | Unidades Administrativas (Reitoria, Pró-Reitorias, Secretarias e Prefeitura Universitária) |
      Elaboração: <span>Nivaldo S. do Rêgo Jr. - CGO/SEPLAN/UFCG</span>
    </div>
  </div>
</div>

<script>
const D = {payload};
let selUGR = null;
let selNDs = new Set();
const expanded = {{}};

document.getElementById('dataAtual').textContent = D.atualizado;

// ── Theme toggle ──────────────────────────────────────────────────────────────
function toggleSidebar() {{
  const body=document.querySelector('.body');
  body.classList.toggle('sb-hidden');
  const hidden=body.classList.contains('sb-hidden');
  document.getElementById('sbToggle').textContent=hidden?'▶':'◀';
  document.getElementById('hamburgerBtn').innerHTML=hidden?'&#9776; Filtros':'&#10005; Filtros';
}}

function toggleTheme() {{
  const html = document.querySelector('html');
  const dark = html.getAttribute('data-theme') === 'dark';
  html.setAttribute('data-theme', dark ? 'light' : 'dark');
  document.getElementById('themeBtn').textContent = dark ? '☀ Modo Claro' : '🌙 Modo Escuro';
}}

// ── Year ──────────────────────────────────────────────────────────────────────
D.anos.slice().reverse().forEach(a => {{
  const o = document.createElement('option');
  o.value = a; o.textContent = a;
  document.getElementById('anoSelect').appendChild(o);
}});
document.getElementById('anoSelect').value = Math.max(...D.anos);

// ── UGR list ──────────────────────────────────────────────────────────────────
function ugrLabel(u) {{
  const m = D.ugr_meta[u];
  if (!m) return u;
  return m.sigla ? m.nome + ' - ' + m.sigla : m.nome;
}}
function buildUGRList(filter='') {{
  const ul = document.getElementById('ugrList');
  ul.innerHTML = '';
  D.ugrs.filter(u => !filter ||
    ugrLabel(u).toLowerCase().includes(filter.toLowerCase()) ||
    u.toLowerCase().includes(filter.toLowerCase())
  ).forEach(u => {{
    const li = document.createElement('li');
    li.className = selUGR === u ? 'active' : '';
    const r = document.createElement('input');
    r.type='radio'; r.name='ugr'; r.checked = selUGR === u;
    li.appendChild(r);
    li.appendChild(document.createTextNode(ugrLabel(u)));
    li.onclick = () => {{ selUGR = selUGR === u ? null : u; buildUGRList(filter); update(); document.getElementById('btnLimparUGR').style.display=selUGR?'':'none'; }};
    ul.appendChild(li);
  }});
}}
function filterUGR(v) {{ buildUGRList(v); }}
buildUGRList();

// ── ND list ───────────────────────────────────────────────────────────────────
function buildNDList(filter='') {{
  const ul = document.getElementById('ndList');
  ul.innerHTML = '';
  D.nds.filter(nd => !filter || nd.toLowerCase().includes(filter.toLowerCase()))
       .forEach(nd => {{
    const li = document.createElement('li');
    li.className = selNDs.has(nd) ? 'active' : '';
    const chk = document.createElement('input');
    chk.type='checkbox'; chk.checked = selNDs.has(nd);
    chk.onchange = e => {{ e.stopPropagation(); selNDs.has(nd)?selNDs.delete(nd):selNDs.add(nd); buildNDList(filter); update(); }};
    li.appendChild(chk);
    li.appendChild(document.createTextNode(nd));
    li.onclick = e => {{ if(e.target===chk) return; selNDs.has(nd)?selNDs.delete(nd):selNDs.add(nd); buildNDList(filter); update(); }};
    ul.appendChild(li);
  }});
}}
function filterND(v) {{ buildNDList(v); }}
function clearND() {{ selNDs.clear(); buildNDList(); update(); }}
function clearUGR() {{ selUGR=null; document.getElementById('btnLimparUGR').style.display='none'; buildUGRList(); update(); }}
buildNDList();

// ── Formatters ────────────────────────────────────────────────────────────────
function fmt(v) {{
  if(!v) return 'R$ -';
  const abs = Math.abs(v);
  if(abs >= 1e6) return 'R$ '+(v/1e6).toLocaleString('pt-BR',{{minimumFractionDigits:2,maximumFractionDigits:2}})+' Mi';
  return 'R$ '+v.toLocaleString('pt-BR',{{minimumFractionDigits:2,maximumFractionDigits:2}});
}}
function fmtK(v) {{
  if(!v&&v!==0) return 'R$ -';
  return 'R$ '+(v/1e3).toLocaleString('pt-BR',{{minimumFractionDigits:2,maximumFractionDigits:2}})+' Mil';
}}
function fmtMi(v) {{
  if(!v&&v!==0) return 'R$ -';
  if(Math.abs(v)>=1e6) return 'R$ '+(v/1e6).toLocaleString('pt-BR',{{minimumFractionDigits:1,maximumFractionDigits:1}})+' Mi';
  return 'R$ '+(v/1e3).toLocaleString('pt-BR',{{minimumFractionDigits:0,maximumFractionDigits:0}})+' Mil';
}}
function pct(a,b){{ return b?Math.round(a/b*1000)/10:0; }}
function isZero(r){{ return !r.DET && !r.EMP && !r.LIQ && !r.PAG; }}
function badgeCls(p){{ return p>=80?'bg-g':p>=40?'bg-a':'bg-r'; }}

// ── Filter helpers ────────────────────────────────────────────────────────────
function rowMatchesND(nd) {{
  if(selNDs.size === 0) return true;
  return selNDs.has(nd.nd_raw || nd.label);
}}
// Achata os filhos de uma UGR pra uma lista plana de NDs, atravessando o
// agrupamento Gestão/Fora do planejamento quando ele existir.
function allNDs(r) {{
  const out=[];
  (r.children||[]).forEach(top=>{{
    if(top.is_group) out.push(...(top.children||[]));
    else out.push(top);
  }});
  return out;
}}
// Uma UGR "casa" com a busca se o nome dela bate, ou se algum ND/NDD/PI dela bate.
function ugrSearchMatch(u, r, searchQ) {{
  if(!searchQ) return true;
  if(ugrLabel(u).toLowerCase().includes(searchQ)) return true;
  return allNDs(r).some(nd=>!isZero(nd)&&(
    nd.label.toLowerCase().includes(searchQ)||(nd.children||[]).some(ndd=>!isZero(ndd)&&(
      ndd.label.toLowerCase().includes(searchQ)||(ndd.children||[]).some(pi=>!isZero(pi)&&pi.label.toLowerCase().includes(searchQ))
    ))
  ));
}}

// ── Expand / Collapse all ─────────────────────────────────────────────────────
function expandAll() {{
  const yd = D.dados[document.getElementById('anoSelect').value] || {{}};
  D.ugrs.filter(u=>yd[u]).forEach(u => {{
    const uid = 'ugr_'+u.replace(/\\W/g,'_');
    expanded[uid]=true;
    (yd[u].children||[]).forEach(top => {{
      const topId = top.is_group ? uid+'_grp_'+top.label.replace(/\\W/g,'_').slice(0,15) : uid;
      if(top.is_group) expanded[topId]=true;
      (top.is_group ? (top.children||[]) : [top]).forEach(nd => {{
        const ndid=topId+'_'+nd.label.replace(/\\W/g,'_').slice(0,25);
        expanded[ndid]=true;
        (nd.children||[]).forEach(ndd => {{ expanded[ndid+'_'+ndd.label.replace(/\\W/g,'_').slice(0,20)]=true; }});
      }});
    }});
  }});
  update();
}}
function collapseAll() {{ Object.keys(expanded).forEach(k=>delete expanded[k]); update(); }}

// Clique simples no ⊞ da UGR: abre a UGR e os grupos PLANEJADO/Fora do
// planejamento (mostrando as NDs), sem descer até as NDDs.
function toggleUgr(uid) {{
  const wasExpanded = !!expanded[uid];
  expanded[uid] = !wasExpanded;
  if(!wasExpanded) {{
    const yd = D.dados[document.getElementById('anoSelect').value] || {{}};
    const u = D.ugrs.find(x => 'ugr_'+x.replace(/\\W/g,'_') === uid);
    if(u && yd[u]) {{
      (yd[u].children||[]).forEach(top => {{
        if(top.is_group) {{
          const topId = uid+'_grp_'+top.label.replace(/\\W/g,'_').slice(0,15);
          expanded[topId] = true;
        }}
      }});
    }}
  }}
  update();
}}

// Duplo-clique no ⊞ da UGR expande tudo dentro dela (grupos, NDs).
function expandUgrAll(uid, evt) {{
  if(evt) evt.stopPropagation();
  const yd = D.dados[document.getElementById('anoSelect').value] || {{}};
  const u = D.ugrs.find(x => 'ugr_'+x.replace(/\\W/g,'_') === uid);
  if(!u || !yd[u]) return;
  expanded[uid]=true;
  (yd[u].children||[]).forEach(top => {{
    const topId = top.is_group ? uid+'_grp_'+top.label.replace(/\\W/g,'_').slice(0,15) : uid;
    if(top.is_group) expanded[topId]=true;
    (top.is_group ? (top.children||[]) : [top]).forEach(nd => {{
      const ndid=topId+'_'+nd.label.replace(/\\W/g,'_').slice(0,25);
      expanded[ndid]=true;
    }});
  }});
  update();
}}

// ── Export CSV ────────────────────────────────────────────────────────────────
function exportCSV() {{
  const ano = document.getElementById('anoSelect').value;
  const yd = D.dados[ano] || {{}};
  const ugrs = selUGR?(yd[selUGR]?[selUGR]:[]):D.ugrs.filter(u=>yd[u]);
  const rows=[['Nivel','Descricao','Detalhado','Empenhado','Liquidado','Pagos','Saldo Disponivel']];
  ugrs.forEach(u => {{
    const r=yd[u]; if(!r) return;
    rows.push(['UGR',ugrLabel(u),r.DET,r.EMP,r.LIQ,r.PAG,r.DISP]);
    allNDs(r).forEach(nd => {{
      if(!rowMatchesND(nd)||isZero(nd)) return;
      rows.push(['ND',nd.label,nd.DET,nd.EMP,nd.LIQ,nd.PAG,nd.DISP]);
      (nd.children||[]).forEach(ndd => {{
        if(isZero(ndd)) return;
        rows.push(['NDD',ndd.label,ndd.DET,ndd.EMP,ndd.LIQ,ndd.PAG,ndd.DISP]);
        (ndd.children||[]).forEach(pi => {{
          if(!isZero(pi)) rows.push(['PI',pi.label,pi.DET,pi.EMP,pi.LIQ,pi.PAG,pi.DISP]);
        }});
      }});
    }});
  }});
  const csv=rows.map(r=>r.map(v=>'"'+String(v).replace(/"/g,'""')+'"').join(',')).join('\\n');
  const a=document.createElement('a');
  a.href=URL.createObjectURL(new Blob(['﻿'+csv],{{type:'text/csv;charset=utf-8'}}));
  a.download='execucao_ufcg_'+ano+'.csv'; a.click();
}}

// ── Export Excel ─────────────────────────────────────────────────────────────
function exportExcel() {{
  const ano = document.getElementById('anoSelect').value;
  const yd = D.dados[ano] || {{}};
  const ugrs = selUGR?(yd[selUGR]?[selUGR]:[]):D.ugrs.filter(u=>yd[u]);
  const rows = [['Nível','Descrição','Detalhado (R$)','Empenhado (R$)','Liquidado (R$)','Pagos (R$)','Saldo Disponível (R$)','% Liquidado']];
  ugrs.forEach(u => {{
    const r=yd[u]; if(!r) return;
    rows.push(['UGR', ugrLabel(u), r.DET, r.EMP, r.LIQ, r.PAG, r.DISP, pct(r.LIQ,r.DET)/100]);
    allNDs(r).forEach(nd => {{
      if(!rowMatchesND(nd)||isZero(nd)) return;
      rows.push(['ND', nd.label, nd.DET, nd.EMP, nd.LIQ, nd.PAG, nd.DISP, pct(nd.LIQ,nd.DET)/100]);
      (nd.children||[]).forEach(ndd => {{
        if(isZero(ndd)) return;
        rows.push(['NDD', ndd.label, ndd.DET, ndd.EMP, ndd.LIQ, ndd.PAG, ndd.DISP, pct(ndd.LIQ,ndd.DET)/100]);
        (ndd.children||[]).forEach(pi => {{
          if(!isZero(pi)) rows.push(['PI', pi.label, pi.DET, pi.EMP, pi.LIQ, pi.PAG, pi.DISP, pct(pi.LIQ,pi.DET)/100]);
        }});
      }});
    }});
  }});
  const ws = XLSX.utils.aoa_to_sheet(rows);
  // Largura das colunas
  ws['!cols'] = [{{wch:6}},{{wch:60}},{{wch:18}},{{wch:18}},{{wch:18}},{{wch:18}},{{wch:20}},{{wch:14}}];
  // Formatar colunas monetárias e percentual
  const fmt_brl = 'R$ #,##0.00';
  const fmt_pct = '0.0%';
  for(let i=1;i<rows.length;i++) {{
    ['C','D','E','F','G'].forEach(col => {{
      const cell = ws[col+(i+1)];
      if(cell) cell.z = fmt_brl;
    }});
    const pcell = ws['H'+(i+1)];
    if(pcell) pcell.z = fmt_pct;
  }}
  // Cabeçalho em negrito
  ['A1','B1','C1','D1','E1','F1','G1','H1'].forEach(ref => {{
    if(ws[ref]) ws[ref].s = {{font:{{bold:true}}}};
  }});
  const wb = XLSX.utils.book_new();
  XLSX.utils.book_append_sheet(wb, ws, 'Execução '+ano);
  // Aba de resumo KPI
  const kpiData = [
    ['Indicador','Valor (R$)','% do Dotado'],
    ['Dotado', parseFloat(document.getElementById('kDET').textContent.replace(/[^0-9,]/g,'').replace(',','.'))*1000 || 0, 1],
    ['Empenhado', 0, parseFloat((document.getElementById('kEMPs').textContent||'0'))/100],
    ['Liquidado', 0, parseFloat((document.getElementById('kLIQs').textContent||'0'))/100],
    ['Saldo Disponível', 0, null]
  ];
  const ws2 = XLSX.utils.aoa_to_sheet(kpiData);
  ws2['!cols']=[{{wch:20}},{{wch:18}},{{wch:14}}];
  XLSX.utils.book_append_sheet(wb, ws2, 'Resumo');
  XLSX.writeFile(wb, 'Execucao_Orcamentaria_UFCG_'+ano+'.xlsx');
}}

// ── Export PDF ────────────────────────────────────────────────────────────────
function exportPDF() {{
  // Expandir tudo antes de imprimir para mostrar o detalhamento completo
  expandAll();
  setTimeout(() => {{
    window.print();
  }}, 300);
}}

// ── Copy link ─────────────────────────────────────────────────────────────────
function copyLink(btn) {{
  const ano=document.getElementById('anoSelect').value;
  const h='ano='+ano+(selUGR?'&ugr='+encodeURIComponent(selUGR):'')+
    ([...selNDs].length?'&nd='+ [...selNDs].map(encodeURIComponent).join(','):'');
  navigator.clipboard.writeText(location.href.split('#')[0]+'#'+h).then(()=>{{
    const orig=btn.innerHTML; btn.textContent='✓ Copiado!';
    setTimeout(()=>{{btn.innerHTML=orig;}},2000);
  }});
}}

// ── Restore URL hash ──────────────────────────────────────────────────────────
(function(){{
  const h=location.hash.slice(1); if(!h) return;
  const p=new URLSearchParams(h);
  if(p.get('ano')) document.getElementById('anoSelect').value=p.get('ano');
  if(p.get('ugr')) selUGR=p.get('ugr');
  if(p.get('nd')) p.get('nd').split(',').forEach(n=>selNDs.add(decodeURIComponent(n)));
  buildUGRList(); buildNDList();
}})();

// ── Main update ───────────────────────────────────────────────────────────────
function update() {{
  const ano=document.getElementById('anoSelect').value;
  const yd=D.dados[ano]||{{}};
  const ugrs=selUGR?(yd[selUGR]?[selUGR]:[]):D.ugrs.filter(u=>yd[u]);
  const searchQ=(document.getElementById('tblSearch').value||'').trim().toLowerCase();

  let tDET=0,tEMP=0,tLIQ=0,tPAG=0,tLIMITE=0;
  ugrs.forEach(u=>{{
    const r=yd[u]; if(!r) return;
    if(!ugrSearchMatch(u,r,searchQ)) return;
    if(selNDs.size===0){{ tDET+=r.DET;tEMP+=r.EMP;tLIQ+=r.LIQ;tPAG+=r.PAG; if(r.LIMITE) tLIMITE+=r.LIMITE; }}
    else allNDs(r).forEach(nd=>{{
      if(!rowMatchesND(nd)) return;
      tDET+=nd.DET;tEMP+=nd.EMP;tLIQ+=nd.LIQ;tPAG+=nd.PAG;
    }});
  }});

  // KPI cards
  document.getElementById('kDET').textContent=fmtMi(tDET);
  document.getElementById('kEMP').textContent=fmtMi(tEMP);
  document.getElementById('kEMPs').textContent=pct(tEMP,tDET)+'% do dotado';
  document.getElementById('kLIQ').textContent=fmtMi(tLIQ);
  document.getElementById('kLIQs').textContent=pct(tLIQ,tDET)+'% do dotado';
  document.getElementById('kDISP').textContent=fmtMi(tDET-tLIQ);

  // Bars
  const pEMP=pct(tEMP,tDET),pLIQ=pct(tLIQ,tDET),pPAG=pct(tPAG,tDET);
  document.getElementById('bDET').textContent=fmtK(tDET);
  document.getElementById('bEMP').style.width=Math.max(pEMP,12)+'%';
  document.getElementById('bEMP').textContent=fmtK(tEMP)+' ('+pEMP+'%)';
  document.getElementById('bLIQ').style.width=Math.max(pLIQ,12)+'%';
  document.getElementById('bLIQ').textContent=fmtK(tLIQ)+' ('+pLIQ+'%)';
  document.getElementById('bPAG').style.width=Math.max(pPAG,8)+'%';
  document.getElementById('bPAG').textContent=fmtK(tPAG)+' ('+pPAG+'%)';

  // Table
  const tbody=document.getElementById('tbody');
  tbody.innerHTML='';

  ugrs.forEach(u=>{{
    const r=yd[u]; if(!r) return;
    let uDET=r.DET,uEMP=r.EMP,uLIQ=r.LIQ,uPAG=r.PAG,uDISP=r.DISP;
    if(selNDs.size>0){{
      uDET=0;uEMP=0;uLIQ=0;uPAG=0;uDISP=0;
      (r.children||[]).forEach(nd=>{{
        if(!rowMatchesND(nd)) return;
        uDET+=nd.DET;uEMP+=nd.EMP;uLIQ+=nd.LIQ;uPAG+=nd.PAG;uDISP+=nd.DISP;
      }});
      if(!uDET&&!uEMP&&!uLIQ&&!uPAG) return;
    }}
    const uid='ugr_'+u.replace(/\\W/g,'_');
    const exp=!!expanded[uid];
    const lbl=ugrLabel(u);
    const uPct=pct(uLIQ,uDET);
    const badge='<span class="exec-badge '+badgeCls(uPct)+'">'+uPct+'% liq.</span>';
    let limiteCell='<td class="td-limite">—</td>';
    if(!exp && selNDs.size===0 && r.LIMITE && r.GESTAO){{
      limiteCell='<td class="td-limite">'+fmt(r.LIMITE)+'</td>';
    }}
    const ugrMatch=!searchQ||lbl.toLowerCase().includes(searchQ);
    const childNDMatch=searchQ?allNDs(r).some(nd=>rowMatchesND(nd)&&!isZero(nd)&&(
      nd.label.toLowerCase().includes(searchQ)||(nd.children||[]).some(ndd=>!isZero(ndd)&&(
        ndd.label.toLowerCase().includes(searchQ)||(ndd.children||[]).some(pi=>!isZero(pi)&&pi.label.toLowerCase().includes(searchQ))
      ))
    )):false;
    const showUGR=ugrMatch||childNDMatch||exp;
    const trU=document.createElement('tr');
    trU.className='ugr-row';
    const numCells = '<td class="td-det">'+fmt(uDET)+'</td><td class="td-emp">'+fmt(uEMP)+'</td><td class="td-liq">'+fmt(uLIQ)+'</td><td class="td-pag">'+fmt(uPAG)+'</td><td>'+fmt(uDISP)+'</td>';
    trU.innerHTML='<td><span class="expand-btn" onclick="toggleUgr(\\''+uid+'\\')" ondblclick="expandUgrAll(\\''+uid+'\\',event)">'+(exp?'⊟':'⊞')+'</span>'+lbl+badge+'</td>'
      +limiteCell+numCells;
    if(!showUGR) trU.classList.add('tr-hidden');
    tbody.appendChild(trU);
    if(!exp&&!childNDMatch) return;

    (r.children||[]).forEach(top=>{{
      const isGroup=!!top.is_group;
      if(isGroup && isZero(top)) return;
      const groupNDs=isGroup?(top.children||[]):[top];
      const topChildMatch=searchQ?groupNDs.some(nd=>rowMatchesND(nd)&&!isZero(nd)&&(
        nd.label.toLowerCase().includes(searchQ)||(nd.children||[]).some(ndd=>!isZero(ndd)&&(
          ndd.label.toLowerCase().includes(searchQ)||(ndd.children||[]).some(pi=>!isZero(pi)&&pi.label.toLowerCase().includes(searchQ))
        ))
      )):false;

      let topId=uid, ndIndentClass='indent1', nddIndentClass='indent2';
      if(isGroup){{
        topId=uid+'_grp_'+top.label.replace(/\\W/g,'_').slice(0,15);
        const topExp=!!expanded[topId];
        const topMatch=!searchQ||top.label.toLowerCase().includes(searchQ)||ugrMatch;
        let limCell='<td class="td-limite">—</td>';
        let excedTag='';
        if(r.LIMITE && top.label.indexOf('PLANEJADO')===0){{
          limCell='<td class="td-limite">'+fmt(r.LIMITE)+'</td>';
          if(top.DET > r.LIMITE){{
            const excedente = top.DET - r.LIMITE;
            excedTag=' <span class="excedente-tag">&#9888; Excedido em '+fmt(excedente)+'</span>';
          }}
        }}
        const trG=document.createElement('tr');
        trG.className='group-row'+(top.label.indexOf('Fora')===0?' group-row-fora':' group-row-gestao');
        if(!topMatch&&!topChildMatch&&!topExp) trG.classList.add('tr-hidden');
        trG.innerHTML='<td class="indent1"><span class="expand-btn" onclick="toggle(\\''+topId+'\\',this)">'+(topExp?'⊟':'⊞')+'</span>↳ '+top.label+excedTag+'</td>'
          +limCell+'<td class="td-det">'+fmt(top.DET)+'</td>'
          +'<td class="td-emp">'+fmt(top.EMP)+'</td>'
          +'<td class="td-liq">'+fmt(top.LIQ)+'</td><td class="td-pag">'+fmt(top.PAG)+'</td>'
          +'<td>'+fmt(top.DISP)+'</td>';
        tbody.appendChild(trG);
        if(!topExp&&!topChildMatch) return;
        ndIndentClass='indent2'; nddIndentClass='indent3';
      }}

      groupNDs.forEach(nd=>{{
        if(!rowMatchesND(nd)||isZero(nd)) return;
        const ndMatch=!searchQ||nd.label.toLowerCase().includes(searchQ)||ugrMatch;
        const ndChildMatch=searchQ&&!ndMatch?(nd.children||[]).some(ndd=>!isZero(ndd)&&(
          ndd.label.toLowerCase().includes(searchQ)||(ndd.children||[]).some(pi=>!isZero(pi)&&pi.label.toLowerCase().includes(searchQ))
        )):false;
        const ndid=topId+'_'+nd.label.replace(/\\W/g,'_').slice(0,25);
        const ndExp=!!expanded[ndid];
        const trN=document.createElement('tr');
        trN.className='nd-row';
        if(!ndMatch&&!ndChildMatch&&!ndExp) trN.classList.add('tr-hidden');
        trN.innerHTML='<td class="'+ndIndentClass+'"><span class="expand-btn" onclick="toggle(\\''+ndid+'\\',this)">'+(ndExp?'⊟':'⊞')+'</span>'+nd.label+'</td>'
          +'<td></td>'
          +'<td class="td-det">'+fmt(nd.DET)+'</td><td class="td-emp">'+fmt(nd.EMP)+'</td>'
          +'<td class="td-liq">'+fmt(nd.LIQ)+'</td><td class="td-pag">'+fmt(nd.PAG)+'</td>'
          +'<td>'+fmt(nd.DISP)+'</td>';
        tbody.appendChild(trN);
        if(!ndExp&&!ndChildMatch) return;

        (nd.children||[]).forEach(ndd=>{{
          if(isZero(ndd)) return;
          const nddMatch=!searchQ||ndd.label.toLowerCase().includes(searchQ)||ndMatch;
          const nddPiMatch=searchQ&&!nddMatch?(ndd.children||[]).some(pi=>!isZero(pi)&&pi.label.toLowerCase().includes(searchQ)):false;
          const trD=document.createElement('tr');
          trD.className='ndd-row';
          if(!nddMatch&&!nddPiMatch) trD.classList.add('tr-hidden');
          const piRows=(ndd.children||[]).filter(pi=>!isZero(pi));
          const tipHtml=piRows.length?'<span class="pi-tip">'
            +piRows.map(pi=>'· '+pi.label.split(' ')[0]).join('  ')
            +'</span>':'';
          trD.innerHTML='<td class="'+nddIndentClass+'"><div class="ndd-cell">'
            +'<span style="display:inline-block;width:21px"></span>'
            +ndd.label+tipHtml+'</div></td>'
            +'<td></td>'
            +'<td class="td-det">'+fmt(ndd.DET)+'</td><td class="td-emp">'+fmt(ndd.EMP)+'</td>'
            +'<td class="td-liq">'+fmt(ndd.LIQ)+'</td><td class="td-pag">'+fmt(ndd.PAG)+'</td>'
            +'<td>'+fmt(ndd.DISP)+'</td>';
          tbody.appendChild(trD);
        }});
      }});
    }});
  }});

  let tLimiteCell='<td class="td-limite">—</td>';
  if(selNDs.size===0 && tLIMITE){{
    tLimiteCell='<td class="td-limite">'+fmt(tLIMITE)+'</td>';
  }}
  const ttr=document.createElement('tr');
  ttr.className='total-row';
  ttr.innerHTML='<td>Total</td>'
    +tLimiteCell
    +'<td class="td-det">'+fmt(tDET)+'</td><td class="td-emp">'+fmt(tEMP)+'</td>'
    +'<td class="td-liq">'+fmt(tLIQ)+'</td><td class="td-pag">'+fmt(tPAG)+'</td>'
    +'<td>'+fmt(tDET-tLIQ)+'</td>';
  tbody.appendChild(ttr);
}}

function toggle(id,btn) {{ expanded[id]=!expanded[id]; update(); }}

update();
</script>
</body>
</html>"""

with open(HTML_OUT,'w',encoding='utf-8') as f:
    f.write(HTML)

# Publicar no GitHub Pages
import shutil, subprocess

IS_CI = bool(os.getenv('CI'))
if IS_CI:
    # No GitHub Actions o próprio checkout já configura um token com permissão de push
    SITE_DIR = BASE
    shutil.copy(HTML_OUT, os.path.join(SITE_DIR, 'index.html'))
    subprocess.run(['git', '-C', SITE_DIR, 'add', 'index.html'], check=True)
    msg = f'Atualizacao automatica {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    result = subprocess.run(['git', '-C', SITE_DIR, 'commit', '-m', msg])
    if result.returncode == 0:
        subprocess.run(['git', '-C', SITE_DIR, 'push'], check=True)
        print('Painel publicado!')
    else:
        print('Nenhuma alteracao para publicar (HTML identico ao anterior).')
else:
    SITE_DIR = r'C:\Users\SAMSUNG\dev\painel-orcamentario-ufcg'
    TOKEN    = os.getenv('GITHUB_TOKEN')
    REMOTE   = f'https://{TOKEN}@github.com/seplan-ufcg/painel-orcamentario-ufcg.git'

    shutil.copy(HTML_OUT, os.path.join(SITE_DIR, 'index.html'))
    subprocess.run(['git', '-C', SITE_DIR, 'add', 'index.html'], check=True)
    msg = f'Atualizacao automatica {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    result = subprocess.run(['git', '-C', SITE_DIR, 'commit', '-m', msg])
    if result.returncode == 0:
        subprocess.run(['git', '-C', SITE_DIR, 'push', REMOTE, 'main'], check=True)
        print('Painel publicado em https://seplan-ufcg.github.io/painel-orcamentario-ufcg/')
    else:
        print('Nenhuma alteracao para publicar (HTML identico ao anterior).')
print(f"Painel gerado: {HTML_OUT}")
